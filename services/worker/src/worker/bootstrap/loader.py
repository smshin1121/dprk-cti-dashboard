"""Read the v1.0 Bootstrap workbook into schema-ready dicts.

This module is intentionally thin: it opens the ``.xlsx`` via openpyxl,
maps each sheet's column headers to the snake_case field names that the
pydantic schemas in :mod:`worker.bootstrap.schemas` expect, and yields
one dict per data row.

Validation, normalization, and database writes all happen later in the
pipeline. Keeping the loader dumb means a column drift in the real
workbook fails loudly at the schema layer (``extra="forbid"`` in the
pydantic models) rather than being silently tolerated here.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, Mapping, Sequence

from openpyxl import load_workbook


__all__ = [
    "WorkbookRow",
    "WorkbookLoader",
    "SHEET_HEADER_MAP",
]


# Maps each sheet name to the ordered tuple of (workbook header, schema field).
# The loader uses this to translate raw cells into pydantic-ready dicts.
# Sheet names and header casings must match the v1.0 workbook exactly.
SHEET_HEADER_MAP: Mapping[str, Sequence[tuple[str, str]]] = {
    "Actors": (
        ("Name", "name"),
        ("Named by", "named_by"),
        ("Associated Group", "associated_group"),
        ("First seen", "first_seen"),
        ("Last seen", "last_seen"),
    ),
    "Reports": (
        ("Published", "published"),
        ("Author", "author"),
        ("Title", "title"),
        ("URL", "url"),
        ("Tags", "tags"),
    ),
    "Incidents": (
        ("Reported", "reported"),
        ("Victims", "victims"),
        ("Motivations", "motivations"),
        ("Sectors", "sectors"),
        ("Countries", "countries"),
    ),
}


class WorkbookLoaderError(ValueError):
    """Raised when the workbook's shape does not match v1.0 expectations."""


@dataclass(frozen=True)
class WorkbookRow:
    """One data row pulled from the workbook.

    ``sheet`` is the sheet's v1.0 name (``"Actors"``, ``"Reports"``,
    ``"Incidents"``). ``index`` is 1-based within the sheet (row 1 is
    the header; the first data row is ``index=1``). ``data`` is the
    header-to-cell mapping translated to schema field names.
    """

    sheet: str
    index: int
    data: dict[str, object]


class WorkbookLoader:
    """Iterate a v1.0 bootstrap workbook one row at a time."""

    def __init__(self, path: Path | str) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"workbook not found: {self.path}")

    def iter_rows(self, sheet: str) -> Iterator[WorkbookRow]:
        """Yield every data row in ``sheet`` as a :class:`WorkbookRow`.

        Header-only sheets yield nothing. Blank rows (every cell None)
        are skipped so trailing empty rows in the workbook do not
        become schema errors.
        """
        if sheet not in SHEET_HEADER_MAP:
            raise WorkbookLoaderError(
                f"unknown sheet {sheet!r}; expected one of "
                f"{tuple(SHEET_HEADER_MAP.keys())}"
            )

        wb = load_workbook(self.path, data_only=True, read_only=True)
        try:
            if sheet not in wb.sheetnames:
                raise WorkbookLoaderError(
                    f"sheet {sheet!r} not in workbook; actual sheets: {wb.sheetnames}"
                )
            ws = wb[sheet]
            expected = SHEET_HEADER_MAP[sheet]
            expected_headers = [h for h, _ in expected]
            expected_fields = [f for _, f in expected]

            row_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(row_iter)
            except StopIteration:
                return

            header_tuple = tuple(header_row[: len(expected_headers)])
            if list(header_tuple) != expected_headers:
                raise WorkbookLoaderError(
                    f"sheet {sheet!r} header mismatch; "
                    f"expected {expected_headers!r} but got {list(header_tuple)!r}"
                )

            data_index = 0
            for raw in row_iter:
                cells = tuple(raw[: len(expected_fields)])
                if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in cells):
                    continue
                data_index += 1
                mapped: dict[str, object] = {
                    field: cell for field, cell in zip(expected_fields, cells)
                }
                yield WorkbookRow(sheet=sheet, index=data_index, data=mapped)
        finally:
            wb.close()

    def iter_all(self) -> Iterator[WorkbookRow]:
        """Yield every data row across every expected sheet in order."""
        for sheet in SHEET_HEADER_MAP:
            yield from self.iter_rows(sheet)
