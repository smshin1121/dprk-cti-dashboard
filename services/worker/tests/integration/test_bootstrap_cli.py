"""End-to-end tests for worker.bootstrap.cli.run_bootstrap.

These drive the full pipeline against the committed fixture and
against small synthetic workbooks built per-test. The db_session
fixture from ``services/worker/tests/conftest.py`` provides a fresh
sqlite in-memory schema per test, and ``run_bootstrap`` is called
directly (not through subprocess / asyncio.run) so assertions can
inspect the returned :class:`ExitDecision` and the resulting DB
state.

The scenarios cover the PR #6 review brief:
  - --dry-run persists nothing
  - --limit is a global cap across all sheets, not per-sheet
  - Dead-letter JSONL is not created on a clean run
  - Dead-letter JSONL carries sheet, row_index, raw_payload,
    error_class, and message for every failure
  - The three exit-code branches of D5 are each exercised end-to-end
  - Idempotency: a second run over the same inputs is a no-op
"""

from __future__ import annotations

import datetime as dt
import io
import json
from pathlib import Path
from typing import Sequence

import pytest
import sqlalchemy as sa
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from worker.bootstrap.cli import run_bootstrap
from worker.bootstrap.errors import ExitCode
from worker.bootstrap.tables import (
    codenames_table,
    groups_table,
    incident_countries_table,
    incident_motivations_table,
    incident_sectors_table,
    incidents_table,
    reports_table,
    sources_table,
)


REPO_ROOT = Path(__file__).resolve().parents[4]
FIXTURE = REPO_ROOT / "services/worker/tests/fixtures/bootstrap_sample.xlsx"
ALIASES = REPO_ROOT / "data/dictionaries/aliases.yml"


# ---------------------------------------------------------------------------
# Synthetic workbook builder
# ---------------------------------------------------------------------------


def _write_synthetic_workbook(
    path: Path,
    *,
    actors: Sequence[dict] = (),
    reports: Sequence[dict] = (),
    incidents: Sequence[dict] = (),
) -> Path:
    """Build a minimal v1.0-shaped workbook with the given rows.

    Empty sheet lists still produce a header row so the loader's
    strict header check stays happy.
    """
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_a = wb.create_sheet("Actors")
    ws_a.append(["Name", "Named by", "Associated Group", "First seen", "Last seen"])
    for row in actors:
        ws_a.append([
            row.get("name"),
            row.get("named_by"),
            row.get("associated_group"),
            row.get("first_seen"),
            row.get("last_seen"),
        ])

    ws_r = wb.create_sheet("Reports")
    ws_r.append(["Published", "Author", "Title", "URL", "Tags"])
    for row in reports:
        ws_r.append([
            row.get("published"),
            row.get("author"),
            row.get("title"),
            row.get("url"),
            row.get("tags"),
        ])

    ws_i = wb.create_sheet("Incidents")
    ws_i.append(["Reported", "Victims", "Motivations", "Sectors", "Countries"])
    for row in incidents:
        ws_i.append([
            row.get("reported"),
            row.get("victims"),
            row.get("motivations"),
            row.get("sectors"),
            row.get("countries"),
        ])

    wb.save(path)
    return path


async def _count(session: AsyncSession, table) -> int:
    result = await session.execute(sa.select(sa.func.count()).select_from(table))
    return result.scalar_one()


# ---------------------------------------------------------------------------
# Exit-code branches (D5)
# ---------------------------------------------------------------------------


async def test_run_bootstrap_curated_happy_fixture_is_clean_exit(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """All-happy workbook -> exit 0 clean, zero failures."""
    wb_path = _write_synthetic_workbook(
        tmp_path / "happy.xlsx",
        actors=[
            {
                "name": "Lazarus Group",
                "named_by": "Kaspersky",
                "associated_group": "Lazarus",
                "first_seen": dt.date(2009, 2, 1),
                "last_seen": dt.date(2025, 12, 15),
            },
            {
                "name": "Kimsuky",
                "named_by": "Kaspersky",
                "associated_group": "Kimsuky",
                "first_seen": dt.date(2012, 9, 10),
                "last_seen": dt.date(2025, 10, 1),
            },
        ],
        reports=[
            {
                "published": dt.date(2024, 3, 15),
                "author": "Mandiant",
                "title": "Lazarus returns with new macOS backdoor",
                "url": "https://example.com/r1",
                "tags": "#lazarus #crypto",
            },
        ],
        incidents=[
            {
                "reported": dt.date(2022, 3, 23),
                "victims": "Ronin Network",
                "motivations": "financial",
                "sectors": "crypto",
                "countries": "VN",
            },
        ],
    )

    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=tmp_path / "errors.jsonl",
        dry_run=False,
        limit=None,
        stdout=stdout,
    )

    assert decision.code == ExitCode.OK
    assert decision.total == 4
    assert decision.failures == 0
    assert "0 failures" in decision.summary
    assert await _count(db_session, reports_table) == 1
    assert await _count(db_session, incidents_table) == 1
    # No dead-letter file on a clean run.
    assert not (tmp_path / "errors.jsonl").exists()


async def test_run_bootstrap_below_5pct_is_warning_exit_zero(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """21 valid rows + 1 invalid = 4.76% failure rate, under the 5%
    threshold. The exit code is still 0 but the summary must read
    "tolerance" so the operator can tell a clean run from a
    warning run."""
    actors = [
        {
            "name": f"Actor {i}",
            "associated_group": "Lazarus",
            "first_seen": dt.date(2020, 1, 1),
            "last_seen": dt.date(2024, 1, 1),
        }
        for i in range(21)
    ]
    # 22nd row is invalid — empty name.
    actors.append(
        {
            "name": "",
            "associated_group": "Lazarus",
            "first_seen": dt.date(2024, 1, 1),
            "last_seen": dt.date(2024, 1, 1),
        }
    )

    wb_path = _write_synthetic_workbook(tmp_path / "warning.xlsx", actors=actors)

    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=tmp_path / "errors.jsonl",
        dry_run=False,
        limit=None,
        stdout=stdout,
    )

    assert decision.total == 22
    assert decision.failures == 1
    assert decision.failure_rate == pytest.approx(1 / 22)
    assert decision.code == ExitCode.OK
    assert "tolerance" in decision.summary
    # Dead-letter file WAS created for the 1 failure.
    assert (tmp_path / "errors.jsonl").exists()


async def test_run_bootstrap_fixture_trips_above_5pct(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """The committed stress fixture has enough failure_case rows to
    exceed the 5% threshold. The exit code must be non-zero."""
    stdout = io.StringIO()
    errors_path = tmp_path / "errors.jsonl"

    decision = await run_bootstrap(
        db_session,
        workbook=FIXTURE,
        aliases_path=ALIASES,
        errors_path=errors_path,
        dry_run=False,
        limit=None,
        stdout=stdout,
    )

    assert decision.code == ExitCode.THRESHOLD_EXCEEDED
    assert decision.total == 32  # 10 actors + 12 reports + 10 incidents
    assert decision.failures > 0
    assert decision.failure_rate > 0.05
    assert "exceeds" in decision.summary
    # Dead-letter file was created.
    assert errors_path.exists()


# ---------------------------------------------------------------------------
# --dry-run persists nothing
# ---------------------------------------------------------------------------


async def test_run_bootstrap_dry_run_writes_no_rows(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """--dry-run must roll back the transaction, so the target
    tables stay at zero rows even though validation and
    normalization ran over every input."""
    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=FIXTURE,
        aliases_path=ALIASES,
        errors_path=None,
        dry_run=True,
        limit=None,
        stdout=stdout,
    )

    # Row counts in every destination table are zero.
    assert await _count(db_session, groups_table) == 0
    assert await _count(db_session, codenames_table) == 0
    assert await _count(db_session, sources_table) == 0
    assert await _count(db_session, reports_table) == 0
    assert await _count(db_session, incidents_table) == 0
    assert await _count(db_session, incident_motivations_table) == 0
    assert await _count(db_session, incident_sectors_table) == 0
    assert await _count(db_session, incident_countries_table) == 0

    # Dry-run exit code still reflects the data-quality gate.
    assert decision.total == 32
    assert decision.failures > 0
    # Stdout announces dry-run.
    assert "dry-run" in stdout.getvalue()


# ---------------------------------------------------------------------------
# --limit is a global cap across sheets
# ---------------------------------------------------------------------------


async def test_run_bootstrap_limit_is_global_across_sheets(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """5 actors + 5 reports + 5 incidents = 15 rows available.
    --limit=7 should stop after 7 total rows (not 7 per sheet), and
    must honor the sheet-declaration order (Actors, Reports, then
    Incidents). So the expected state after the run is: 5 actors,
    2 reports, 0 incidents."""
    actors = [
        {
            "name": f"Codename{i}",
            "associated_group": "Lazarus",
        }
        for i in range(5)
    ]
    reports = [
        {
            "published": dt.date(2024, 3, 15),
            "author": "Mandiant",
            "title": f"Report {i}",
            "url": f"https://example.com/r{i}",
            "tags": "#crypto",  # sector tag only; does not create a codename
        }
        for i in range(5)
    ]
    incidents = [
        {
            "reported": dt.date(2024, 1, 1),
            "victims": f"Incident{i}",
            "motivations": "financial",
            "sectors": "crypto",
            "countries": "US",
        }
        for i in range(5)
    ]
    wb_path = _write_synthetic_workbook(
        tmp_path / "limit.xlsx",
        actors=actors,
        reports=reports,
        incidents=incidents,
    )

    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=None,
        dry_run=False,
        limit=7,
        stdout=stdout,
    )

    assert decision.total == 7
    assert decision.failures == 0

    assert await _count(db_session, codenames_table) == 5
    assert await _count(db_session, reports_table) == 2
    assert await _count(db_session, incidents_table) == 0


async def test_run_bootstrap_limit_rejects_nonpositive(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    stdout = io.StringIO()
    with pytest.raises(ValueError):
        await run_bootstrap(
            db_session,
            workbook=FIXTURE,
            aliases_path=ALIASES,
            errors_path=None,
            dry_run=False,
            limit=0,
            stdout=stdout,
        )


# ---------------------------------------------------------------------------
# Dead-letter JSONL schema
# ---------------------------------------------------------------------------


async def test_dead_letter_not_created_on_clean_run(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    wb_path = _write_synthetic_workbook(
        tmp_path / "clean.xlsx",
        actors=[
            {"name": "Lazarus Group", "associated_group": "Lazarus"},
        ],
    )
    errors_path = tmp_path / "nested" / "errors.jsonl"

    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=errors_path,
        dry_run=False,
        limit=None,
        stdout=io.StringIO(),
    )
    assert decision.failures == 0
    assert not errors_path.exists()


async def test_dead_letter_contains_all_required_fields(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """Every dead-letter line must carry sheet, row_index,
    raw_payload, error_class, and message so an operator can
    reproduce the failure from the committed fixture."""
    wb_path = _write_synthetic_workbook(
        tmp_path / "failing.xlsx",
        actors=[
            {"name": "Lazarus Group", "associated_group": "Lazarus"},
        ],
        reports=[
            {
                "published": dt.date(2024, 3, 15),
                "author": "Mandiant",
                "title": "",  # forces RowValidationError
                "url": "https://example.com/x",
                "tags": "#lazarus",
            },
        ],
        incidents=[
            {
                "reported": dt.date(2024, 1, 1),
                "victims": "Corp",
                "countries": "XX",  # invalid ISO code
            },
        ],
    )
    errors_path = tmp_path / "errors.jsonl"

    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=errors_path,
        dry_run=False,
        limit=None,
        stdout=io.StringIO(),
    )

    assert decision.failures == 2
    assert errors_path.exists()

    lines = errors_path.read_text(encoding="utf-8").splitlines()
    assert len(lines) == 2
    entries = [json.loads(line) for line in lines]

    required = {"sheet", "row_index", "raw_payload", "error_class", "message"}
    for entry in entries:
        assert required.issubset(entry.keys())

    # The Reports failure came in first (Reports sheet is processed
    # second, after Actors which had no failures).
    assert entries[0]["sheet"] == "Reports"
    assert entries[0]["error_class"] == "ValidationError"  # pydantic v2
    assert "title" in entries[0]["message"].lower()

    # The Incidents failure (invalid country) comes next.
    assert entries[1]["sheet"] == "Incidents"
    assert entries[1]["raw_payload"]["countries"] == "XX"


# ---------------------------------------------------------------------------
# Idempotency — the PR #5 guarantee still holds end-to-end
# ---------------------------------------------------------------------------


async def test_run_bootstrap_second_run_is_noop_on_happy_fixture(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """Two consecutive non-dry-run passes over a happy fixture must
    produce identical row counts — the PR #5 idempotency contract
    survives the full CLI wrapper."""
    wb_path = _write_synthetic_workbook(
        tmp_path / "happy.xlsx",
        actors=[
            {"name": "Lazarus Group", "associated_group": "Lazarus"},
            {"name": "Kimsuky", "associated_group": "Kimsuky"},
        ],
        reports=[
            {
                "published": dt.date(2024, 3, 15),
                "author": "Mandiant",
                "title": "r1",
                "url": "https://example.com/r1",
                "tags": "#lazarus",
            },
            {
                "published": dt.date(2024, 4, 1),
                "author": "Kaspersky",
                "title": "r2",
                "url": "https://example.com/r2",
                "tags": "#kimsuky",
            },
        ],
        incidents=[
            {
                "reported": dt.date(2024, 1, 1),
                "victims": "Corp",
                "motivations": "financial",
                "sectors": "crypto",
                "countries": "US",
            },
        ],
    )

    async def _snapshot() -> dict[str, int]:
        return {
            "groups": await _count(db_session, groups_table),
            "codenames": await _count(db_session, codenames_table),
            "sources": await _count(db_session, sources_table),
            "reports": await _count(db_session, reports_table),
            "incidents": await _count(db_session, incidents_table),
        }

    await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=None,
        dry_run=False,
        limit=None,
        stdout=io.StringIO(),
    )
    first = await _snapshot()
    assert first["reports"] == 2
    assert first["incidents"] == 1

    await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=None,
        dry_run=False,
        limit=None,
        stdout=io.StringIO(),
    )
    second = await _snapshot()
    # The exact count of codenames depends on which tags create
    # them — what matters for the idempotency contract is that the
    # second pass is a no-op relative to the first.
    assert second == first


# ---------------------------------------------------------------------------
# --errors-path="" disables the writer entirely
# ---------------------------------------------------------------------------


async def test_run_bootstrap_preserves_caller_outer_transaction(
    db_session: AsyncSession, tmp_path: Path, aliases=None
) -> None:
    """Codex round 3 P1: if the caller is already inside a
    transaction when run_bootstrap is called, the ETL must NOT
    commit or roll back the caller's outer transaction. It is
    allowed to manipulate only its own SAVEPOINT.

    Setup: write one row through the caller's transaction BEFORE
    calling run_bootstrap. That row must survive the ETL (both
    dry-run and real) because the caller hasn't committed yet.
    """
    # Insert a sentinel row directly through the caller's session
    # inside the caller's transaction. This starts the session's
    # autobegin transaction.
    await db_session.execute(
        sa.insert(groups_table).values(name="SentinelGroup")
    )
    # Prove the insert landed inside the caller's uncommitted txn.
    assert await _count(db_session, groups_table) == 1

    wb_path = _write_synthetic_workbook(
        tmp_path / "happy.xlsx",
        actors=[{"name": "Lazarus Group", "associated_group": "Lazarus"}],
    )

    # Dry-run: run_bootstrap must roll back only its own SAVEPOINT;
    # the sentinel row must still be visible through the caller's
    # session.
    await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=None,
        dry_run=True,
        limit=None,
        stdout=io.StringIO(),
    )
    assert await _count(db_session, groups_table) == 1  # sentinel survives

    # Non-dry-run: run_bootstrap releases its SAVEPOINT (its own
    # rows become part of the caller's transaction) but must NOT
    # commit the caller's outer transaction. Both the sentinel and
    # the ETL's Lazarus group are visible; neither is yet committed
    # from the caller's perspective.
    await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=None,
        dry_run=False,
        limit=None,
        stdout=io.StringIO(),
    )
    assert await _count(db_session, groups_table) == 2  # sentinel + Lazarus
    # The caller's outer transaction is still active — run_bootstrap
    # did not commit it out from under the caller.
    assert db_session.in_transaction()


async def test_run_bootstrap_errors_path_none_disables_writer(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """Passing ``errors_path=None`` to run_bootstrap means no file
    is ever touched even when failures occur. Useful for CI
    invocations that don't want artifacts."""
    decision = await run_bootstrap(
        db_session,
        workbook=FIXTURE,
        aliases_path=ALIASES,
        errors_path=None,
        dry_run=False,
        limit=None,
        stdout=io.StringIO(),
    )
    assert decision.failures > 0
    # No files under tmp_path.
    assert list(tmp_path.iterdir()) == []
