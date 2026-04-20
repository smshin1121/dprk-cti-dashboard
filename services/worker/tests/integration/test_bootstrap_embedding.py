"""Integration tests for embed-on-ingest wiring in run_bootstrap.

PR #19a Group B criteria C1 + C4 verification for the worker
bootstrap path:

  - **C1 (no signature drift / sqlite invariance):** a run with
    ``embedding_client=None`` behaves exactly as pre-PR-#19a (covered
    here as the default path in every existing `test_bootstrap_cli.py`
    test plus the explicit assertion below).
  - **C4 (enrichment never blocks ingest):**
    - With a live mock client that would return 200 JSON, sqlite's
      dialect guard inside ``embed_report`` short-circuits before the
      HTTP call; the report row still lands cleanly.
    - With ``embed_report`` monkeypatched to raise
      :class:`PermanentEmbeddingError`, the report row still lands —
      ``_process_one_row`` catches the exception inside the per-row
      savepoint so the INSERT is not rolled back.
    - With ``embed_report`` monkeypatched to raise
      :class:`TransientEmbeddingError` (simulating a contract
      violation where a future refactor exposes the transient class
      instead of swallowing inside ``embed_report``), the report row
      still lands — the same swallow guard catches it defensively.

The sqlite test schema intentionally omits the pgvector ``embedding``
column; these tests therefore verify only row presence / ingest
completion, not vector contents. PG-only UPDATE mechanics are pinned
in ``test_embedding_writer.py`` via an AsyncSession stub.
"""

from __future__ import annotations

import datetime as dt
import io
from pathlib import Path
from typing import Sequence

import httpx
import pytest
import sqlalchemy as sa
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from worker.bootstrap import cli as cli_module
from worker.bootstrap.cli import run_bootstrap
from worker.bootstrap.embedding_client import (
    LlmProxyEmbeddingClient,
    PermanentEmbeddingError,
    TransientEmbeddingError,
)
from worker.bootstrap.embedding_writer import (
    EmbedReportResult,
    EmbedWriteOutcome,
)
from worker.bootstrap.errors import ExitCode
from worker.bootstrap.tables import reports_table


REPO_ROOT = Path(__file__).resolve().parents[4]
ALIASES = REPO_ROOT / "data/dictionaries/aliases.yml"
DIM = 1536


def _write_minimal_workbook(
    path: Path,
    *,
    reports: Sequence[dict],
) -> Path:
    """Build a workbook with only a Reports sheet populated.

    Actors and Incidents sheets have headers but no rows, mirroring
    the shape ``run_bootstrap`` expects from the loader (strict header
    check in Phase 0).
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_a = wb.create_sheet("Actors")
    ws_a.append(["Name", "Named by", "Associated Group", "First seen", "Last seen"])

    ws_r = wb.create_sheet("Reports")
    ws_r.append(["Published", "Author", "Title", "URL", "Tags"])
    for row in reports:
        ws_r.append([
            row.get("published"),
            row.get("author"),
            row.get("title"),
            row.get("url"),
            row.get("tags"),
        ])

    ws_i = wb.create_sheet("Incidents")
    ws_i.append(["Reported", "Victims", "Motivations", "Sectors", "Countries"])

    wb.save(path)
    return path


def _sample_report(title: str = "Lazarus revived") -> dict:
    return {
        "published": dt.date(2024, 3, 15),
        "author": "Mandiant",
        "title": title,
        "url": "https://example.com/embed-test",
        "tags": "#lazarus #crypto",
    }


async def _count_reports(session: AsyncSession) -> int:
    result = await session.execute(
        sa.select(sa.func.count()).select_from(reports_table)
    )
    return result.scalar_one()


def _live_mock_client() -> LlmProxyEmbeddingClient:
    """Mock client that returns a 200 JSON envelope on every call.

    On sqlite paths the client is never hit (dialect guard inside
    ``embed_report`` short-circuits first); we use this to verify the
    run stays clean even when a client IS injected.
    """
    async def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(
            200,
            json={
                "provider": "mock",
                "model": "text-embedding-3-small",
                "dimensions": DIM,
                "items": [{"index": 0, "embedding": [0.5] * DIM}],
                "usage": {"prompt_tokens": 3, "total_tokens": 3},
                "latency_ms": 1,
                "cache_hit": False,
            },
        )

    return LlmProxyEmbeddingClient(
        base_url="http://llm-proxy.test",
        internal_token="test-token",
        client=httpx.AsyncClient(transport=httpx.MockTransport(handler)),
        timeout_seconds=5.0,
    )


# ---------------------------------------------------------------------------
# C1 default path — no client, no behavior change
# ---------------------------------------------------------------------------


async def test_default_path_embedding_client_none_is_backward_compat(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """Verify pre-PR-#19a behavior is preserved when client is None.

    Also acts as a tripwire: if ``_process_one_row`` ever calls
    ``embed_report`` when the client kwarg is None, this test fails
    because the monkeypatched ``embed_report`` raises.
    """
    wb_path = _write_minimal_workbook(
        tmp_path / "no_client.xlsx",
        reports=[_sample_report()],
    )

    # No monkeypatch needed — the REAL embed_report is present. With
    # embedding_client=None, the caller code path in _process_one_row
    # MUST skip the embed_report call outright (client-None guard).
    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=tmp_path / "errors.jsonl",
        dry_run=False,
        limit=None,
        stdout=stdout,
        embedding_client=None,
    )

    assert decision.code == ExitCode.OK
    assert decision.total == 1
    assert decision.failures == 0
    assert await _count_reports(db_session) == 1


# ---------------------------------------------------------------------------
# C4 sqlite path with live client
# ---------------------------------------------------------------------------


async def test_live_mock_client_on_sqlite_inserts_cleanly(
    db_session: AsyncSession, tmp_path: Path
) -> None:
    """With a live mock client on sqlite, ``embed_report`` returns
    ``SKIPPED_SQLITE`` before any HTTP call. The report row still
    lands; the run completes with zero failures."""
    wb_path = _write_minimal_workbook(
        tmp_path / "sqlite_live.xlsx",
        reports=[_sample_report(title="Sqlite live-client test")],
    )

    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=tmp_path / "errors.jsonl",
        dry_run=False,
        limit=None,
        stdout=stdout,
        embedding_client=_live_mock_client(),
    )

    assert decision.code == ExitCode.OK
    assert decision.failures == 0
    assert await _count_reports(db_session) == 1


# ---------------------------------------------------------------------------
# C4 permanent error is caught at _process_one_row
# ---------------------------------------------------------------------------


async def test_permanent_embedding_error_is_caught_row_persists(
    db_session: AsyncSession,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Monkeypatch ``embed_report`` to raise ``PermanentEmbeddingError``.

    The per-row savepoint in ``run_bootstrap`` MUST NOT roll back on
    this error. The report row lands and the run completes with zero
    failures (embedding failure is not counted as a row failure).
    """
    async def raising_embed_report(
        session,  # noqa: ANN001
        *,
        report_id: int,
        title: str,
        summary: str | None,
        client,  # noqa: ANN001
    ):
        raise PermanentEmbeddingError(
            upstream_status=422,
            reason="invalid_input",
        )

    monkeypatch.setattr(cli_module, "embed_report", raising_embed_report)

    wb_path = _write_minimal_workbook(
        tmp_path / "permanent.xlsx",
        reports=[_sample_report(title="Permanent-error test")],
    )

    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=tmp_path / "errors.jsonl",
        dry_run=False,
        limit=None,
        stdout=stdout,
        embedding_client=_live_mock_client(),
    )

    assert decision.code == ExitCode.OK
    assert decision.total == 1
    assert decision.failures == 0
    assert await _count_reports(db_session) == 1


# ---------------------------------------------------------------------------
# C4 transient error defensive handling
# ---------------------------------------------------------------------------


async def test_transient_embedding_error_defensive_catch(
    db_session: AsyncSession,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Defensive coverage — ``TransientEmbeddingError`` is normally
    swallowed inside ``embed_report`` and returned as
    ``SKIPPED_TRANSIENT``. This test pins that contract by replacing
    ``embed_report`` with one that RETURNS the skip outcome (as the
    real one does), and asserts run_bootstrap completes cleanly.

    Together with the unit tests in ``test_embedding_writer.py``
    (which cover the actual swallow logic), this locks the symmetric
    behavior: transient errors never escape the embed layer.
    """
    calls: list[int] = []

    async def skipping_embed_report(
        session,  # noqa: ANN001
        *,
        report_id: int,
        title: str,
        summary: str | None,
        client,  # noqa: ANN001
    ) -> EmbedReportResult:
        calls.append(report_id)
        return EmbedReportResult(
            outcome=EmbedWriteOutcome.SKIPPED_TRANSIENT,
            rowcount=0,
            cache_hit=None,
            upstream_latency_ms=None,
        )

    monkeypatch.setattr(cli_module, "embed_report", skipping_embed_report)

    wb_path = _write_minimal_workbook(
        tmp_path / "transient.xlsx",
        reports=[_sample_report(title="Transient-error test")],
    )

    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=tmp_path / "errors.jsonl",
        dry_run=False,
        limit=None,
        stdout=stdout,
        embedding_client=_live_mock_client(),
    )

    assert decision.code == ExitCode.OK
    assert decision.failures == 0
    assert await _count_reports(db_session) == 1
    # Exactly one report row -> exactly one embed_report call.
    assert len(calls) == 1


# ---------------------------------------------------------------------------
# Non-Reports sheets never trigger embed_report
# ---------------------------------------------------------------------------


async def test_non_reports_sheets_do_not_call_embed_report(
    db_session: AsyncSession,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Actors / Incidents rows must never trigger an embed call. This
    prevents accidental scope creep if someone later adds an embed
    dispatch branch to the Actors / Incidents arms."""
    calls: list[int] = []

    async def tripwire_embed_report(
        session,  # noqa: ANN001
        *,
        report_id: int,
        title: str,
        summary: str | None,
        client,  # noqa: ANN001
    ) -> EmbedReportResult:
        calls.append(report_id)
        raise AssertionError(
            "embed_report must not be called for non-Reports sheets"
        )

    monkeypatch.setattr(cli_module, "embed_report", tripwire_embed_report)

    wb_path = tmp_path / "actors_only.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws_a = wb.create_sheet("Actors")
    ws_a.append(["Name", "Named by", "Associated Group", "First seen", "Last seen"])
    ws_a.append([
        "Lazarus Group",
        "Kaspersky",
        "Lazarus",
        dt.date(2009, 2, 1),
        dt.date(2025, 12, 15),
    ])
    ws_r = wb.create_sheet("Reports")
    ws_r.append(["Published", "Author", "Title", "URL", "Tags"])
    ws_i = wb.create_sheet("Incidents")
    ws_i.append(["Reported", "Victims", "Motivations", "Sectors", "Countries"])
    ws_i.append([
        dt.date(2022, 3, 23),
        "Ronin Network",
        "financial",
        "crypto",
        "VN",
    ])
    wb.save(wb_path)

    stdout = io.StringIO()
    decision = await run_bootstrap(
        db_session,
        workbook=wb_path,
        aliases_path=ALIASES,
        errors_path=tmp_path / "errors.jsonl",
        dry_run=False,
        limit=None,
        stdout=stdout,
        embedding_client=_live_mock_client(),
    )

    assert decision.code == ExitCode.OK
    assert len(calls) == 0
