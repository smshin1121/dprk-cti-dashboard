"""Unit tests for scripts/generate_bootstrap_fixture.py.

The generator lives in ``scripts/`` (not on sys.path as a package),
so these tests load it via :mod:`importlib.util` from the repo root.
The test file covers the new ``strip_failure_cases`` filter added
for the data-quality-tests CI job — without regressing the default
"regenerate stress fixture" behaviour that worker-tests relies on
for its drift check.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path
from types import ModuleType
from typing import Any

import pytest
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# importlib loader — generator is in scripts/, not a package
# ---------------------------------------------------------------------------


REPO_ROOT = Path(__file__).resolve().parents[4]
GENERATOR_PATH = REPO_ROOT / "scripts" / "generate_bootstrap_fixture.py"


def _load_generator() -> ModuleType:
    spec = importlib.util.spec_from_file_location(
        "generate_bootstrap_fixture", GENERATOR_PATH
    )
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


@pytest.fixture(scope="module")
def generator() -> ModuleType:
    return _load_generator()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _minimal_source() -> dict[str, list[dict[str, Any]]]:
    """Three-sheet shape with one happy row and one failure_case row
    per sheet. Enough to exercise both the preserve and drop paths
    without depending on the committed YAML."""
    return {
        "actors": [
            {
                "name": "Lazarus Group",
                "associated_group": "Lazarus",
                "_tag": "happy",
            },
            {
                "name": "",
                "associated_group": "Lazarus",
                "_tag": "failure_case",
                "_reason": "empty name",
            },
        ],
        "reports": [
            {
                "title": "Lazarus returns",
                "url": "https://example.com/r1",
                "_tag": "happy",
            },
            {
                "title": "Bad Report",
                "url": "not-a-url",
                "_tag": "failure_case",
                "_reason": "malformed URL",
            },
        ],
        "incidents": [
            {
                "victims": "Ronin Network",
                "countries": "VN",
                "_tag": "happy",
            },
            {
                "victims": "",
                "countries": "XX",
                "_tag": "failure_case",
                "_reason": "invalid country",
            },
        ],
    }


# ---------------------------------------------------------------------------
# strip_failure_cases — pure function
# ---------------------------------------------------------------------------


class TestStripFailureCases:
    def test_removes_every_failure_case_row(
        self, generator: ModuleType
    ) -> None:
        filtered = generator.strip_failure_cases(_minimal_source())
        for sheet_key in ("actors", "reports", "incidents"):
            tags = [row.get("_tag") for row in filtered[sheet_key]]
            assert "failure_case" not in tags
            assert len(filtered[sheet_key]) == 1  # only the happy row

    def test_preserves_rows_with_no_tag(self, generator: ModuleType) -> None:
        """Untagged rows count as happy — they must survive the filter."""
        source = {
            "actors": [{"name": "A"}, {"name": "B", "_tag": "failure_case"}],
            "reports": [],
            "incidents": [],
        }
        filtered = generator.strip_failure_cases(source)
        assert filtered["actors"] == [{"name": "A"}]

    def test_preserves_rows_with_other_tag_values(
        self, generator: ModuleType
    ) -> None:
        source = {
            "actors": [
                {"name": "A", "_tag": "happy"},
                {"name": "B", "_tag": "edge_case"},
                {"name": "C", "_tag": "failure_case"},
            ],
            "reports": [],
            "incidents": [],
        }
        filtered = generator.strip_failure_cases(source)
        assert [row["name"] for row in filtered["actors"]] == ["A", "B"]

    def test_does_not_mutate_source(self, generator: ModuleType) -> None:
        source = _minimal_source()
        before_lens = {k: len(v) for k, v in source.items()}
        _ = generator.strip_failure_cases(source)
        after_lens = {k: len(v) for k, v in source.items()}
        assert before_lens == after_lens
        assert all(len(v) == 2 for v in source.values())

    def test_non_list_sheet_value_passes_through(
        self, generator: ModuleType
    ) -> None:
        """If a sheet key maps to something other than a list, the
        filter leaves it alone so :func:`build_workbook` can raise
        its own validation error later."""
        source = {
            "actors": "not a list",  # type: ignore[dict-item]
            "reports": [{"_tag": "failure_case"}],
            "incidents": [],
        }
        filtered = generator.strip_failure_cases(source)
        assert filtered["actors"] == "not a list"
        assert filtered["reports"] == []

    def test_empty_source_is_preserved(self, generator: ModuleType) -> None:
        source: dict[str, list[dict[str, Any]]] = {
            "actors": [],
            "reports": [],
            "incidents": [],
        }
        filtered = generator.strip_failure_cases(source)
        assert filtered == source
        assert filtered is not source  # new dict, not aliased


# ---------------------------------------------------------------------------
# CLI — exercises main() and the --strip-failure-cases flag
# ---------------------------------------------------------------------------


class TestMainCli:
    def test_default_run_regenerates_stress_fixture_byte_equivalent(
        self, generator: ModuleType, tmp_path: Path
    ) -> None:
        """No-flag run must produce a workbook that matches the
        committed stress fixture cell-for-cell so the worker-tests
        drift check stays green."""
        out = tmp_path / "stress.xlsx"
        exit_code = generator.main(["--output", str(out)])
        assert exit_code == 0

        committed = load_workbook(
            REPO_ROOT / "services/worker/tests/fixtures/bootstrap_sample.xlsx",
            data_only=True,
        )
        regen = load_workbook(out, data_only=True)
        assert committed.sheetnames == regen.sheetnames
        for sheet in committed.sheetnames:
            c_rows = [
                tuple(cell.value for cell in row)
                for row in committed[sheet].iter_rows()
            ]
            r_rows = [
                tuple(cell.value for cell in row)
                for row in regen[sheet].iter_rows()
            ]
            assert c_rows == r_rows, f"cell drift in sheet {sheet!r}"

    def test_strip_flag_drops_failure_case_rows(
        self, generator: ModuleType, tmp_path: Path
    ) -> None:
        stress_out = tmp_path / "stress.xlsx"
        happy_out = tmp_path / "happy.xlsx"
        assert generator.main(["--output", str(stress_out)]) == 0
        assert (
            generator.main(
                ["--strip-failure-cases", "--output", str(happy_out)]
            )
            == 0
        )

        stress_wb = load_workbook(stress_out, data_only=True)
        happy_wb = load_workbook(happy_out, data_only=True)

        # Happy variant must have strictly fewer rows in every sheet
        # that contained a failure_case in the source YAML, and must
        # never have MORE rows than the stress variant.
        for sheet in stress_wb.sheetnames:
            stress_count = stress_wb[sheet].max_row
            happy_count = happy_wb[sheet].max_row
            assert happy_count <= stress_count

        # Total stress - total happy must equal the failure_case count
        # from the source YAML. Reading the yaml directly keeps this
        # assertion robust to future fixture churn.
        import yaml

        with (
            REPO_ROOT
            / "services/worker/tests/fixtures/bootstrap_sample.yaml"
        ).open("r", encoding="utf-8") as handle:
            source = yaml.safe_load(handle)
        expected_drop = sum(
            sum(1 for row in rows if row.get("_tag") == "failure_case")
            for rows in source.values()
        )
        stress_total = sum(
            stress_wb[sheet].max_row - 1 for sheet in stress_wb.sheetnames
        )
        happy_total = sum(
            happy_wb[sheet].max_row - 1 for sheet in happy_wb.sheetnames
        )
        assert stress_total - happy_total == expected_drop
        assert expected_drop > 0  # fixture must still contain failures

    def test_strip_flag_never_writes_private_metadata_cells(
        self, generator: ModuleType, tmp_path: Path
    ) -> None:
        """Underscore-prefixed keys like ``_tag`` / ``_reason`` are
        reviewer metadata and must stay out of the workbook cells
        regardless of whether the filter ran."""
        out = tmp_path / "happy.xlsx"
        generator.main(["--strip-failure-cases", "--output", str(out)])
        wb = load_workbook(out, data_only=True)
        for sheet in wb.sheetnames:
            headers = [cell.value for cell in wb[sheet][1]]
            for header in headers:
                assert header is None or not str(header).startswith("_")
