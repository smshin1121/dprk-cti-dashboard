"""Tests for worker.bootstrap.loader against the committed fixture."""

from __future__ import annotations

from pathlib import Path

import pytest

from worker.bootstrap.loader import (
    SHEET_HEADER_MAP,
    WorkbookLoader,
    WorkbookLoaderError,
)


REPO_ROOT = Path(__file__).resolve().parents[4]
FIXTURE = REPO_ROOT / "services/worker/tests/fixtures/bootstrap_sample.xlsx"


@pytest.fixture
def loader() -> WorkbookLoader:
    return WorkbookLoader(FIXTURE)


# ---------------------------------------------------------------------------
# Construction
# ---------------------------------------------------------------------------


def test_missing_file_rejected(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        WorkbookLoader(tmp_path / "nonexistent.xlsx")


# ---------------------------------------------------------------------------
# iter_rows — happy path against committed fixture
# ---------------------------------------------------------------------------


def test_actors_sheet_row_count(loader: WorkbookLoader) -> None:
    rows = list(loader.iter_rows("Actors"))
    assert len(rows) == 10


def test_reports_sheet_row_count(loader: WorkbookLoader) -> None:
    rows = list(loader.iter_rows("Reports"))
    assert len(rows) == 12


def test_incidents_sheet_row_count(loader: WorkbookLoader) -> None:
    rows = list(loader.iter_rows("Incidents"))
    assert len(rows) == 10


def test_first_actor_row_fields(loader: WorkbookLoader) -> None:
    rows = list(loader.iter_rows("Actors"))
    first = rows[0]
    assert first.sheet == "Actors"
    assert first.index == 1
    assert first.data["name"] == "Lazarus Group"
    assert first.data["associated_group"] == "Lazarus"
    # openpyxl returns dates as datetime; the pydantic schemas coerce
    # them later, but here we just confirm the mapping preserved them.
    assert first.data["first_seen"] is not None


def test_row_data_uses_schema_field_names(loader: WorkbookLoader) -> None:
    """The loader's job is to translate workbook headers (e.g.
    "Associated Group") into snake_case field names the pydantic
    schemas consume. Verify the mapping holds."""
    rows = list(loader.iter_rows("Reports"))
    first = rows[0]
    expected_fields = {"published", "author", "title", "url", "tags"}
    assert set(first.data.keys()) == expected_fields


def test_rows_have_incrementing_index(loader: WorkbookLoader) -> None:
    rows = list(loader.iter_rows("Actors"))
    indices = [row.index for row in rows]
    assert indices == list(range(1, 11))


# ---------------------------------------------------------------------------
# iter_all — cross-sheet iteration
# ---------------------------------------------------------------------------


def test_iter_all_yields_every_row(loader: WorkbookLoader) -> None:
    rows = list(loader.iter_all())
    assert len(rows) == 10 + 12 + 10


def test_iter_all_sheet_ordering(loader: WorkbookLoader) -> None:
    sheet_order = [row.sheet for row in loader.iter_all()]
    # Actors come first, then Reports, then Incidents — same order as
    # SHEET_HEADER_MAP declares.
    assert sheet_order[:10] == ["Actors"] * 10
    assert sheet_order[10:22] == ["Reports"] * 12
    assert sheet_order[22:] == ["Incidents"] * 10


# ---------------------------------------------------------------------------
# Error cases
# ---------------------------------------------------------------------------


def test_unknown_sheet_rejected(loader: WorkbookLoader) -> None:
    with pytest.raises(WorkbookLoaderError, match="unknown sheet"):
        list(loader.iter_rows("NoSuchSheet"))


def test_sheet_header_map_has_three_sheets() -> None:
    assert set(SHEET_HEADER_MAP.keys()) == {"Actors", "Reports", "Incidents"}
    for _, fields in SHEET_HEADER_MAP.items():
        # Every sheet must have >= 1 mapping.
        assert len(fields) >= 1


# ---------------------------------------------------------------------------
# Schema-drift rejection — the P2 fix Codex flagged
# ---------------------------------------------------------------------------


def _write_minimal_workbook(path: Path, actors_header_row: list, actors_data_row: list) -> None:
    """Build a tiny Actors-only workbook with the given header / data
    rows. Used by the drift-rejection tests."""
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Actors")
    ws.append(actors_header_row)
    ws.append(actors_data_row)
    wb.save(path)


def test_loader_rejects_extra_header_column(tmp_path: Path) -> None:
    """An upstream workbook that adds a new column must fail loud.
    Without this, the loader would silently drop the new column and
    data."""
    bad = tmp_path / "drift.xlsx"
    _write_minimal_workbook(
        bad,
        actors_header_row=[
            "Name",
            "Named by",
            "Associated Group",
            "First seen",
            "Last seen",
            "NEW EXTRA COLUMN",  # upstream drift
        ],
        actors_data_row=[
            "Lazarus",
            "Kaspersky",
            "Lazarus",
            None,
            None,
            "extra cell",
        ],
    )
    loader = WorkbookLoader(bad)
    with pytest.raises(WorkbookLoaderError, match="header mismatch"):
        list(loader.iter_rows("Actors"))


def test_loader_rejects_header_in_wrong_order(tmp_path: Path) -> None:
    bad = tmp_path / "wrong_order.xlsx"
    _write_minimal_workbook(
        bad,
        actors_header_row=[
            "Named by",  # swapped
            "Name",  # swapped
            "Associated Group",
            "First seen",
            "Last seen",
        ],
        actors_data_row=["Kaspersky", "Lazarus", "Lazarus", None, None],
    )
    loader = WorkbookLoader(bad)
    with pytest.raises(WorkbookLoaderError, match="header mismatch"):
        list(loader.iter_rows("Actors"))


def test_loader_rejects_data_row_with_overflow_cell(tmp_path: Path) -> None:
    """If the header is fine but a data row has a populated cell
    beyond the expected column width, that is still schema drift and
    must fail loud."""
    bad = tmp_path / "overflow.xlsx"
    _write_minimal_workbook(
        bad,
        actors_header_row=[
            "Name",
            "Named by",
            "Associated Group",
            "First seen",
            "Last seen",
        ],
        # 6 populated cells — one more than the header declares.
        actors_data_row=["Lazarus", "Kaspersky", "Lazarus", None, None, "rogue"],
    )
    loader = WorkbookLoader(bad)
    with pytest.raises(WorkbookLoaderError, match="populated cell beyond"):
        list(loader.iter_rows("Actors"))


def test_loader_tolerates_trailing_none_in_header(tmp_path: Path) -> None:
    """openpyxl can emit trailing None cells when the worksheet was
    saved with more columns allocated than filled. Those trailing
    Nones must not trip the strict header comparison."""
    from openpyxl import Workbook

    path = tmp_path / "trailing.xlsx"
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Actors")
    ws.append([
        "Name",
        "Named by",
        "Associated Group",
        "First seen",
        "Last seen",
    ])
    # Explicitly write a None into a 6th column so openpyxl allocates
    # the extra cell.
    ws.cell(row=1, column=6, value=None)
    ws.append(["Lazarus", "Kaspersky", "Lazarus", None, None])
    wb.save(path)

    loader = WorkbookLoader(path)
    rows = list(loader.iter_rows("Actors"))
    assert len(rows) == 1
    assert rows[0].data["name"] == "Lazarus"
